"""Export helpers for PortDetector desktop panels."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable


def _normalize_path(output_path: str, suffix: str) -> Path:
    path = Path(output_path)
    if path.suffix.lower() != suffix:
        path = path.with_suffix(suffix)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def export_report_spreadsheet(
    report_rows: list[dict],
    summary_text: str,
    hours: int,
    output_path: str,
) -> str:
    """Export the uptime report to XLSX, with CSV fallback when needed."""
    path = _normalize_path(output_path, ".xlsx")

    try:
        import openpyxl
        from openpyxl.styles import Border, Font, Alignment, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return export_report_csv(report_rows, summary_text, hours, str(path.with_suffix(".csv")))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uptime Report"

    header_font = Font(name="Pretendard", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="2A2A4A"))

    summary_lines = [
        "PortDetector Uptime Report",
        summary_text,
        f"Generated: last {hours} hr",
    ]
    for row_idx, line in enumerate(summary_lines, 1):
        ws.cell(row=row_idx, column=1, value=line)
    ws.append([])

    header_row = 5
    headers = [
        "Device", "IP", "Category", "Importance", "Health", "Uptime %",
        "Up (min)", "Down (min)", "Disconnects", "Avg RTT (ms)",
        "Status Changes", "Why It Matters", "Next Action",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    num_align = Alignment(horizontal="right")

    for row_idx, row_data in enumerate(report_rows, header_row + 1):
        ws.cell(row=row_idx, column=1, value=row_data["name"])
        ws.cell(row=row_idx, column=2, value=row_data["ip"])
        ws.cell(row=row_idx, column=3, value=row_data["category"])
        ws.cell(row=row_idx, column=4, value=row_data["importance"])
        ws.cell(row=row_idx, column=5, value=row_data["report_severity"])

        pct_cell = ws.cell(row=row_idx, column=6, value=row_data["uptime_pct"] / 100)
        pct_cell.number_format = "0.0%"
        pct_cell.alignment = num_align

        for col, key in [(7, "up_mins"), (8, "down_mins")]:
            cell = ws.cell(row=row_idx, column=col, value=row_data[key])
            cell.number_format = "#,##0"
            cell.alignment = num_align

        ws.cell(row=row_idx, column=9, value=row_data["disconnects"]).alignment = num_align

        avg_rtt = row_data["avg_rtt"]
        rtt_cell = ws.cell(row=row_idx, column=10, value=avg_rtt if avg_rtt is not None else "")
        if avg_rtt is not None:
            rtt_cell.number_format = "#,##0.0"
        rtt_cell.alignment = num_align

        ws.cell(row=row_idx, column=11, value=row_data["status_changes"]).alignment = num_align
        ws.cell(row=row_idx, column=12, value=row_data["report_summary"])
        ws.cell(row=row_idx, column=13, value=row_data["report_action"])

    widths = [20, 16, 14, 12, 12, 12, 12, 12, 12, 14, 14, 34, 34]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return str(path)


def export_report_csv(
    report_rows: list[dict],
    summary_text: str,
    hours: int,
    output_path: str,
) -> str:
    """CSV fallback for the uptime report."""
    path = _normalize_path(output_path, ".csv")
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["PortDetector Uptime Report"])
        writer.writerow([summary_text])
        writer.writerow([f"Generated: last {hours} hr"])
        writer.writerow([])
        writer.writerow([
            "Device", "IP", "Category", "Importance", "Health", "Uptime %",
            "Up (min)", "Down (min)", "Disconnects", "Avg RTT (ms)",
            "Status Changes", "Why It Matters", "Next Action",
        ])
        for row_data in report_rows:
            writer.writerow([
                f"{row_data['name']}",
                row_data["ip"],
                row_data["category"],
                row_data["importance"],
                row_data["report_severity"],
                f"{row_data['uptime_pct']:.1f}%",
                row_data["up_mins"],
                row_data["down_mins"],
                row_data["disconnects"],
                "" if row_data["avg_rtt"] is None else row_data["avg_rtt"],
                row_data["status_changes"],
                row_data["report_summary"],
                row_data["report_action"],
            ])
    return str(path)


def export_history_csv(
    history_rows: list[dict],
    output_path: str,
) -> str:
    """Export filtered history rows to CSV."""
    path = _normalize_path(output_path, ".csv")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Timestamp", "Device", "Change", "Severity", "Detail", "RTT"])
        for row in history_rows:
            writer.writerow([
                row.get("timestamp", ""),
                row.get("device", ""),
                row.get("change", ""),
                row.get("severity_label", ""),
                row.get("detail", ""),
                "" if row.get("rtt") is None else row.get("rtt"),
            ])
    return str(path)
