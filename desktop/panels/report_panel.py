"""Uptime Report panel — device availability statistics + Excel export."""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QFileDialog, QMessageBox,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QColor

from desktop.theme import Colors, Fonts
from desktop.i18n import t
from backend.utils.monitoring_presenter import (
    importance_label,
    severity_label,
)


SEVERITY_COLORS = {
    "stable": Colors.CONNECTED,
    "advisory": Colors.WARNING,
    "warning": Colors.WARNING,
    "critical": Colors.DISCONNECTED,
    "emergency": Colors.DISCONNECTED,
}


class ReportPanel(QWidget):
    def __init__(self, uptime_service, config_service, parent=None):
        super().__init__(parent)
        self._uptime_service = uptime_service
        self._config_service = config_service
        self._report_data = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)

        self._title = QLabel(t("report.title"))
        self._title.setStyleSheet(f"font-size: {Fonts.SIZE_XL}px; font-weight: 600; color: {Colors.TEXT}; background: transparent;")
        layout.addWidget(self._title)

        self._guide_label = QLabel(t("guide.report"))
        self._guide_label.setWordWrap(True)
        self._guide_label.setStyleSheet(
            f"font-size: {Fonts.SIZE_XS}px; color: {Colors.TEXT_MUTED}; background: transparent; padding-bottom: 4px;"
        )
        layout.addWidget(self._guide_label)

        # Controls
        ctrl = QHBoxLayout()
        ctrl.setSpacing(8)

        self._period_label = QLabel(t("report.period"))
        self._period_label.setStyleSheet(f"font-size: {Fonts.SIZE_SM}px; color: {Colors.TEXT_DIM}; background: transparent;")
        ctrl.addWidget(self._period_label)

        self._period_combo = QComboBox()
        self._period_combo.addItem(t("report.last_24h"), 24)
        self._period_combo.addItem(t("report.last_12h"), 12)
        self._period_combo.addItem(t("report.last_48h"), 48)
        self._period_combo.addItem(t("report.last_7d"), 168)
        self._period_combo.setFixedWidth(140)
        ctrl.addWidget(self._period_combo)

        self._gen_btn = QPushButton(t("report.generate"))
        self._gen_btn.setObjectName("btn_primary")
        self._gen_btn.setFixedHeight(30)
        self._gen_btn.clicked.connect(self._generate)
        ctrl.addWidget(self._gen_btn)

        self._export_btn = QPushButton(t("report.export_excel"))
        self._export_btn.setFixedHeight(30)
        self._export_btn.clicked.connect(self._export_excel)
        ctrl.addWidget(self._export_btn)

        ctrl.addStretch()
        layout.addLayout(ctrl)

        self._summary_label = QLabel(t("report.generate_prompt"))
        self._summary_label.setWordWrap(True)
        self._summary_label.setStyleSheet(
            f"""
            QLabel {{
                background-color: {Colors.BG_CARD};
                border: 1px solid {Colors.BORDER};
                border-radius: 8px;
                padding: 10px 12px;
                color: {Colors.TEXT_DIM};
                font-size: {Fonts.SIZE_SM}px;
            }}
        """
        )
        layout.addWidget(self._summary_label)

        # Table
        self._table = QTableWidget()
        self._table_headers = [
            t("report.col_device"), t("report.col_importance"), t("report.col_health"),
            t("report.col_uptime"), t("report.col_disconnects"), t("report.col_avg_rtt"),
            t("report.col_changes"), t("report.col_detail"), t("report.col_action"),
        ]
        self._table.setColumnCount(len(self._table_headers))
        self._table.setHorizontalHeaderLabels(self._table_headers)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        for i in [1, 2, 3, 4, 5, 6]:
            self._table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(1, 90)
        self._table.setColumnWidth(2, 90)
        self._table.setColumnWidth(3, 90)
        self._table.setColumnWidth(4, 95)
        self._table.setColumnWidth(5, 105)
        self._table.setColumnWidth(6, 100)
        self._table.setShowGrid(False)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.verticalHeader().setVisible(False)
        self._table.setStyleSheet(f"QTableWidget {{ alternate-background-color: {Colors.BG_ALT}; }}")
        layout.addWidget(self._table, 1)

    def _update_summary(self, hours):
        if not self._report_data:
            self._summary_label.setText(t("report.no_devices"))
            return

        counts = {
            'critical': 0,
            'warning': 0,
            'advisory': 0,
            'stable': 0,
        }
        for row in self._report_data:
            severity = row.get('report_severity', 'stable')
            if severity in counts:
                counts[severity] += 1
            elif severity == 'emergency':
                counts['critical'] += 1

        headline = (
            f"Last {hours} hr: {counts['critical']} critical, {counts['warning']} warning, "
            f"{counts['advisory']} advisory, {counts['stable']} stable device(s)."
        )
        top = self._report_data[0]
        self._summary_label.setText(
            headline + " "
            f"Top review target: {top['name']} ({severity_label(top['report_severity'])}). "
            f"{top['report_action']}"
        )

    def _generate(self):
        hours = self._period_combo.currentData()
        devices = self._config_service.get_devices()
        delay_threshold_ms = self._config_service.get_settings().get('delay_threshold_ms', 200)
        self._report_data = self._uptime_service.generate_report_data(
            devices,
            hours,
            delay_threshold_ms=delay_threshold_ms,
        )
        self._update_summary(hours)

        self._table.setRowCount(0)

        for row_data in self._report_data:
            row = self._table.rowCount()
            self._table.insertRow(row)

            self._table.setItem(row, 0, QTableWidgetItem(f"{row_data['name']} ({row_data['ip']})"))

            self._table.setItem(row, 1, QTableWidgetItem(importance_label(row_data['importance'])))

            severity = row_data['report_severity']
            severity_item = QTableWidgetItem(severity_label(severity))
            severity_item.setForeground(QColor(SEVERITY_COLORS.get(severity, Colors.TEXT_DIM)))
            self._table.setItem(row, 2, severity_item)

            # Uptime %
            pct = row_data['uptime_pct']
            pct_item = QTableWidgetItem(f"{pct:.1f}%")
            pct_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if pct >= 99:
                pct_item.setForeground(QColor(Colors.CONNECTED))
            elif pct >= 95:
                pct_item.setForeground(QColor(Colors.DELAYED))
            else:
                pct_item.setForeground(QColor(Colors.DISCONNECTED))
            self._table.setItem(row, 3, pct_item)

            # Disconnects
            disc = row_data['disconnects']
            disc_item = QTableWidgetItem(str(disc))
            disc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if disc > 0:
                disc_item.setForeground(QColor(Colors.DISCONNECTED))
            self._table.setItem(row, 4, disc_item)

            # Avg RTT
            avg_rtt = row_data['avg_rtt']
            rtt_text = f"{avg_rtt:,.1f}" if avg_rtt is not None else "--"
            rtt_item = QTableWidgetItem(rtt_text)
            rtt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, 5, rtt_item)

            # Status changes
            sc = row_data['status_changes']
            sc_item = QTableWidgetItem(str(sc))
            sc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, 6, sc_item)

            self._table.setItem(row, 7, QTableWidgetItem(row_data['report_summary']))
            self._table.setItem(row, 8, QTableWidgetItem(row_data['report_action']))

    def _export_excel(self):
        if not self._report_data:
            QMessageBox.information(self, t("common.export"), t("report.export_first"))
            return

        path, _ = QFileDialog.getSaveFileName(
            self, t("report.export_excel"), "uptime_report.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Uptime Report"

            # Header style
            header_font = Font(name="Pretendard", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
            thin_border = Border(
                bottom=Side(style='thin', color='2A2A4A'),
            )

            summary_lines = [
                "PortDetector Uptime Report",
                self._summary_label.text(),
                f"Generated: {self._period_combo.currentText()}",
            ]
            for row_idx, line in enumerate(summary_lines, 1):
                ws.cell(row=row_idx, column=1, value=line)
            ws.append([])

            header_row = 5
            headers = ["Device", "IP", "Category", "Importance", "Health", "Uptime %",
                        "Up (min)", "Down (min)", "Disconnects", "Avg RTT (ms)",
                        "Status Changes", "Why It Matters", "Next Action"]

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data
            data_font = Font(name="Pretendard", size=10)
            num_align = Alignment(horizontal='right')

            for i, row_data in enumerate(self._report_data, header_row + 1):
                ws.cell(row=i, column=1, value=row_data['name'])
                ws.cell(row=i, column=2, value=row_data['ip']).font = data_font
                ws.cell(row=i, column=3, value=row_data['category'])
                ws.cell(row=i, column=4, value=importance_label(row_data['importance']))

                health_cell = ws.cell(row=i, column=5, value=severity_label(row_data['report_severity']))
                health_cell.font = Font(
                    name="Pretendard",
                    size=10,
                    color=SEVERITY_COLORS.get(row_data['report_severity'], Colors.TEXT_DIM).replace("#", "").upper(),
                )

                pct_cell = ws.cell(row=i, column=6, value=row_data['uptime_pct'] / 100)
                pct_cell.number_format = '0.0%'
                pct_cell.alignment = num_align
                pct_cell.font = data_font

                # Conditional color for uptime
                pct = row_data['uptime_pct']
                if pct < 95:
                    pct_cell.font = Font(name="Pretendard", size=10, color="EF476F")
                elif pct < 99:
                    pct_cell.font = Font(name="Pretendard", size=10, color="FFD166")
                else:
                    pct_cell.font = Font(name="Pretendard", size=10, color="06D6A0")

                for col, key in [(7, 'up_mins'), (8, 'down_mins')]:
                    c = ws.cell(row=i, column=col, value=row_data[key])
                    c.number_format = '#,##0'
                    c.alignment = num_align
                    c.font = data_font

                ws.cell(row=i, column=9, value=row_data['disconnects']).alignment = num_align

                avg_rtt = row_data['avg_rtt']
                rtt_cell = ws.cell(row=i, column=10, value=avg_rtt if avg_rtt is not None else '')
                if avg_rtt is not None:
                    rtt_cell.number_format = '#,##0.0'
                rtt_cell.alignment = num_align
                rtt_cell.font = data_font

                ws.cell(row=i, column=11, value=row_data['status_changes']).alignment = num_align
                ws.cell(row=i, column=12, value=row_data['report_summary'])
                ws.cell(row=i, column=13, value=row_data['report_action'])

            # Column widths
            widths = [20, 16, 14, 12, 12, 12, 12, 12, 12, 14, 14, 34, 34]
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = w

            ws.freeze_panes = f"A{header_row + 1}"

            wb.save(path)
            QMessageBox.information(self, t("common.export"), t("report.exported", path=path))
        except ImportError:
            QMessageBox.warning(self, t("common.error"), "openpyxl is required for Excel export.\npip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, t("common.error"), str(e))

    def retranslate(self):
        """Update all translatable strings to the current language."""
        self._title.setText(t("report.title"))
        self._guide_label.setText(t("guide.report"))
        self._period_label.setText(t("report.period"))

        # Update period combo items (preserve current selection by data)
        current_data = self._period_combo.currentData()
        self._period_combo.blockSignals(True)
        self._period_combo.clear()
        self._period_combo.addItem(t("report.last_24h"), 24)
        self._period_combo.addItem(t("report.last_12h"), 12)
        self._period_combo.addItem(t("report.last_48h"), 48)
        self._period_combo.addItem(t("report.last_7d"), 168)
        idx = self._period_combo.findData(current_data)
        if idx >= 0:
            self._period_combo.setCurrentIndex(idx)
        self._period_combo.blockSignals(False)

        self._gen_btn.setText(t("report.generate"))
        self._export_btn.setText(t("report.export_excel"))

        # Update summary if no report generated yet
        if not self._report_data:
            self._summary_label.setText(t("report.generate_prompt"))

        # Update table headers
        self._table_headers = [
            t("report.col_device"), t("report.col_importance"), t("report.col_health"),
            t("report.col_uptime"), t("report.col_disconnects"), t("report.col_avg_rtt"),
            t("report.col_changes"), t("report.col_detail"), t("report.col_action"),
        ]
        self._table.setHorizontalHeaderLabels(self._table_headers)
